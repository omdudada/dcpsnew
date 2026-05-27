import sys
from pathlib import Path

import openpyxl


def _safe(v):
    if v is None:
        return ""
    s = str(v)
    s = s.replace("\r", " ").replace("\n", " ")
    return s[:120]


def main():
    if len(sys.argv) < 2:
        print("Usage: python tools/read_excel_ledger.py <path-to-xlsx>")
        return 2

    xlsx_path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    print("sheets:", wb.sheetnames)

    # Prefer the first sheet with "Ledger" content (fallback to first sheet).
    ws = wb[wb.sheetnames[0]]
    print("active:", ws.title, "rows:", ws.max_row, "cols:", ws.max_column)

    # Print first non-empty 40 rows x 18 cols (values)
    print("\n--- TOP (values) ---")
    for r in range(1, min(41, ws.max_row + 1)):
        row = [ws.cell(r, c).value for c in range(1, 19)]
        if any(v not in (None, "") for v in row):
            print(r, [_safe(v) for v in row])

    # Find the header row that contains "Month" and print 20 rows from there.
    header_row = None
    for r in range(1, min(200, ws.max_row + 1)):
        vals = [str(ws.cell(r, c).value or "").strip().lower() for c in range(1, 40)]
        if any("month" == v or "month" in v for v in vals) and any("opening" in v for v in vals):
            header_row = r
            break

    if header_row:
        print("\n--- DETECTED HEADER ROW:", header_row, "---")
        for r in range(header_row, min(header_row + 21, ws.max_row + 1)):
            row = [ws.cell(r, c).value for c in range(1, 19)]
            if any(v not in (None, "") for v in row):
                print(r, [_safe(v) for v in row])

        # Also show formulas in the "Interest" column candidate (look for cell containing "interest")
        interest_col = None
        for c in range(1, 60):
            hv = str(ws.cell(header_row, c).value or "").strip().lower()
            if "interest" in hv:
                interest_col = c
                break
        if interest_col:
            print("\n--- INTEREST COLUMN:", interest_col, "(formulas) ---")
            for r in range(header_row + 1, min(header_row + 16, ws.max_row + 1)):
                cell = ws.cell(r, interest_col)
                if cell.value not in (None, ""):
                    print(f"R{r}C{interest_col}:", _safe(cell.value))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

